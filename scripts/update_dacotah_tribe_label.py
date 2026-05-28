"""
Update IATH Tribes.xlsx rows for the DACOTAH-glo variants.

Rows 325-327 currently map:
  DACOTAH                  → tribe='Sioux',  reservation='FRN'
  DACOTAH OR SIOUX         → tribe='Sioux',  reservation='FRN'
  DACOTAH OR SIOUX NATION  → tribe='Sioux',  reservation='FRN'

Change: set tribe column to 'Dacotah/Sioux Nation' on all three rows.
Reservation stays 'FRN' (the specific band/reservation is still not
resolvable from the GLO name alone). Existing reservation notes are
left untouched.

Adds a TAG-attributed tribe note recording the change for audit.

Idempotent — re-running detects the TAG and skips.

Usage:
    ./venv/bin/python3 scripts/update_dacotah_tribe_label.py            # dry-run
    ./venv/bin/python3 scripts/update_dacotah_tribe_label.py --apply    # write
"""
import argparse
from openpyxl import load_workbook

XLSX = "/Users/cwm6W/Library/CloudStorage/Box-Box/IATH/tribes/Tribes.xlsx"
TAG  = "[2026-05-28 Claude-assisted draft]"

TARGET_GLOS = {"DACOTAH", "DACOTAH OR SIOUX", "DACOTAH OR SIOUX NATION"}
NEW_TRIBE   = "Dacotah/Sioux Nation"
NOTE = (
    f"{TAG} Tribe label updated from 'Sioux' to 'Dacotah/Sioux Nation' "
    f"to align with the label applied to Sioux Scrip Patent (docClass=SS) "
    f"records. BLM's 'DACOTAH' GLO terminology identifies the nation directly; "
    f"specific band/reservation remains unresolved (reservation stays FRN)."
)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(XLSX, data_only=False)
    ws = wb["Sheet1"]
    edits = []

    for row_idx in range(2, ws.max_row + 1):
        glo = (ws.cell(row=row_idx, column=1).value or "").strip().upper()
        if glo not in TARGET_GLOS:
            continue

        tribe_cell       = ws.cell(row=row_idx, column=2)
        tribe_notes_cell = ws.cell(row=row_idx, column=6)
        current_tribe = (tribe_cell.value or "").strip()
        current_notes = (tribe_notes_cell.value or "").strip()

        if TAG in current_notes:
            print(f"row {row_idx} ({glo}): already tagged — skip")
            continue

        if current_tribe != NEW_TRIBE:
            edits.append((row_idx, 2, current_tribe, NEW_TRIBE,
                          f"tribe → {NEW_TRIBE} for {glo!r}"))
        new_notes = f"{current_notes}\n\n{NOTE}".strip() if current_notes else NOTE
        edits.append((row_idx, 6, current_notes, new_notes, f"tribe note for {glo!r}"))

    print(f"Edits queued: {len(edits)}")
    for r_idx, col, old, new, reason in edits:
        print(f"  row {r_idx} col {col} ({reason})")
        print(f"    OLD: {(old or '(empty)')!r}")
        print(f"    NEW: {new[:120]!r}{'...' if len(str(new)) > 120 else ''}")

    if not args.apply:
        print("\nDRY RUN — pass --apply to write.")
        return
    if not edits:
        print("No edits to apply.")
        return

    for row_idx, col, _, new, _ in edits:
        ws.cell(row=row_idx, column=col).value = new
    wb.save(XLSX)
    print(f"\nSaved {len(edits)} edits to {XLSX}")


if __name__ == "__main__":
    main()
