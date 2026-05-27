"""
Build the `tribe_crosswalk` DB table from the IATH-maintained tribes
spreadsheet at /Users/cwm6W/Library/CloudStorage/Box-Box/IATH/tribes/Tribes.xlsx.

The spreadsheet is hand-built by Christian McMillen: for each distinct
glo_tribe_name string seen in the BLM patent catalog, it records the
authoritative tribe + reservation (or 'FRN' = further research needed,
or NULL = no applicable value), plus alternatives and notes.

This script imports the spreadsheet faithfully (all 8 columns) AND derives
a `canonical_reservation` column where the spreadsheet's auth-reservation
text is normalized via CANONICAL_OVERRIDES — handling the case where the
spreadsheet has two spelling variants for the same reservation (e.g. KCA's
'Kiowa, Comanche, Apache Reservation' vs 'Comanche, Kiowa, and Apache
Reservation' → both normalize to the conventional 'Kiowa, Comanche, and
Apache Reservation').

`canonical_reservation` is NULL when the spreadsheet says FRN or no value;
populated only when the spreadsheet has a real reservation name.

Idempotent — DROPs and rebuilds the table each run.

Usage:
    ./venv/bin/python3 scripts/build_tribe_crosswalk.py
    DATABASE_URL="host=127.0.0.1 port=5433 dbname=allotment_research user=appuser password=allotment-app-2026" \\
        ./venv/bin/python3 scripts/build_tribe_crosswalk.py
"""
import os
import sys
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

DB_URL = os.environ.get("DATABASE_URL", "dbname=allotment_research user=cwm6W")
XLSX   = "/Users/cwm6W/Library/CloudStorage/Box-Box/IATH/tribes/Tribes.xlsx"

# Reservation-name canonicalization — applied at xlsx → DB import time.
# When the spreadsheet has multiple spellings of the same federally-
# recognized reservation, map all of them to the conventional name.
CANONICAL_OVERRIDES = {
    "Kiowa, Comanche, Apache Reservation":     "Kiowa, Comanche, and Apache Reservation",
    "Comanche, Kiowa, and Apache Reservation": "Kiowa, Comanche, and Apache Reservation",
}


def normalize_for_canonical(raw):
    """Return canonical reservation name (None if FRN/empty); apply overrides."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "" or s.upper() == "FRN":
        return None
    return CANONICAL_OVERRIDES.get(s, s)


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"missing {XLSX}")

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        glo = r[0]
        if not glo or not str(glo).strip():
            continue
        rows.append((
            str(glo).strip(),                              # glo_name
            str(glo).strip().upper(),                      # glo_name_normalized
            r[1].strip() if isinstance(r[1], str) else r[1],  # authoritative_tribe
            r[2].strip() if isinstance(r[2], str) else r[2],  # authoritative_reservation
            normalize_for_canonical(r[2]),                 # canonical_reservation
            r[3].strip() if isinstance(r[3], str) else r[3],  # alternative_tribe
            r[4].strip() if isinstance(r[4], str) else r[4],  # alternative_reservation
            r[5].strip() if isinstance(r[5], str) else r[5],  # tribe_notes
            r[6].strip() if isinstance(r[6], str) else r[6],  # reservation_notes
            r[7].strip() if isinstance(r[7], str) else r[7],  # general_notes
        ))
    print(f"Loaded {len(rows)} GLO entries from spreadsheet")

    conn = psycopg2.connect(DB_URL)
    cur  = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS tribe_crosswalk")
    cur.execute("""
        CREATE TABLE tribe_crosswalk (
            id serial PRIMARY KEY,
            glo_name text NOT NULL,
            glo_name_normalized text NOT NULL,
            authoritative_tribe text,
            authoritative_reservation text,
            canonical_reservation text,
            alternative_tribe text,
            alternative_reservation text,
            tribe_notes text,
            reservation_notes text,
            general_notes text
        )
    """)
    cur.execute("CREATE INDEX idx_tribe_crosswalk_glo_norm ON tribe_crosswalk (glo_name_normalized)")
    cur.execute("CREATE INDEX idx_tribe_crosswalk_canon_res ON tribe_crosswalk (canonical_reservation)")
    print("Created tribe_crosswalk table + indexes")

    sql = """
        INSERT INTO tribe_crosswalk
            (glo_name, glo_name_normalized, authoritative_tribe,
             authoritative_reservation, canonical_reservation,
             alternative_tribe, alternative_reservation,
             tribe_notes, reservation_notes, general_notes)
        VALUES %s
    """
    psycopg2.extras.execute_values(cur, sql, rows, page_size=500)
    conn.commit()
    print(f"Inserted {len(rows)} rows")
    print()

    # Verification
    cur.execute("SELECT COUNT(*) FROM tribe_crosswalk")
    print(f"Verification: tribe_crosswalk has {cur.fetchone()[0]} rows")

    cur.execute("""
        SELECT COUNT(DISTINCT canonical_reservation)
        FROM tribe_crosswalk
        WHERE canonical_reservation IS NOT NULL
    """)
    print(f"  Distinct canonical reservations: {cur.fetchone()[0]}")

    cur.execute("""
        SELECT COUNT(*) FROM tribe_crosswalk
        WHERE authoritative_reservation IS NOT NULL
          AND canonical_reservation IS NOT NULL
          AND TRIM(authoritative_reservation) != canonical_reservation
    """)
    print(f"  Rows where canonical_reservation differs from authoritative (override applied): {cur.fetchone()[0]}")

    # Show which overrides actually fired
    cur.execute("""
        SELECT authoritative_reservation, canonical_reservation, COUNT(*)
        FROM tribe_crosswalk
        WHERE authoritative_reservation IS NOT NULL
          AND canonical_reservation IS NOT NULL
          AND TRIM(authoritative_reservation) != canonical_reservation
        GROUP BY authoritative_reservation, canonical_reservation
    """)
    print()
    print("Overrides applied:")
    for raw, canon, n in cur.fetchall():
        print(f"  {n:>3}  {raw!r} → {canon!r}")


if __name__ == "__main__":
    main()
