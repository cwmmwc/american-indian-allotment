"""
Import the cancelled-patents research spreadsheet into the database as
the cancelled_patent_research table.

Source: /Users/cwm6W/Library/CloudStorage/OneDrive-UniversityofVirginia/
        Cancelled patents/Cancelled patents_6.11._cwm.xlsx

Steps:
  1. Drop and recreate cancelled_patent_research from create_cancelled_patent_research.sql
  2. Read the xlsx
  3. Normalize each row (strip whitespace, parse dates, handle empty cells)
  4. Insert all rows
  5. Report verification stats: total imported, joined-to-all_patents count,
     and which spreadsheet entries don't have a matching patent record in
     all_patents (potential data-entry typos or missing-from-BLM cases).

Run from project root:
    ./venv/bin/python3 scripts/import_cancelled_patents.py
"""
import os
import sys
from datetime import datetime, date
import openpyxl
import psycopg2
import psycopg2.extras

DB_URL = os.environ.get("DATABASE_URL", "dbname=allotment_research user=cwm6W")
XLSX = "/Users/cwm6W/Library/CloudStorage/OneDrive-UniversityofVirginia/Cancelled patents/Cancelled patents_6.11._cwm.xlsx"
SCHEMA_SQL = "sql/create_cancelled_patent_research.sql"

# Spreadsheet column mapping (0-indexed). Built from the file header inspection.
COL = {
    "name":                   0,   # "Name "
    "allotment_number":       1,
    "tribe_reservation":      2,
    "state":                  3,
    "_blank1":                4,
    "reason_for_cancellation": 5,
    "cancellation_date":      6,
    "fee_patent_date":        7,
    "patent_number":          8,
    "ccf_number":             9,
    "gender":                 10,
    "carlisle_yn":            11,
    "comments":               12,
    "_marker":                13,  # "do not fill in columns to the right"
    "_blank2":                14,
    "in_dtpo":                15,
    "ccf_alt":                16,
}


def s(v):
    """Normalize a string cell: strip, None -> None, empty -> None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return str(v).strip() or None


def _correct_century_bug(d):
    """Excel's 2-digit-year interpretation maps '16' to 2016, but Dawes-era
    forced-fee patents and cancellations are all 19XX. If a date comes in
    with year 2000+, roll it back 100 years. We use 2000 as the threshold
    because (a) the spreadsheet documents 1916-1936 events, and (b) Excel's
    default 2-digit-year cutoff produces 20XX for any value entered without
    a 19 prefix. If a future user genuinely enters a 21st-century date for
    legitimate reasons, this will mis-correct — re-run import after fixing
    the source, or remove the correction here."""
    if d is None:
        return None
    if d.year >= 2000:
        try:
            return d.replace(year=d.year - 100)
        except ValueError:
            # Feb 29 in a leap year that's not a leap century-back; rare
            return d.replace(year=d.year - 100, day=28)
    return d


def d(v):
    """Normalize a date cell: keep datetime/date as-is, parse strings, ignore
    unparseable. Corrects Excel's 2-digit-year century bug (2016 -> 1916)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return _correct_century_bug(v.date())
    if isinstance(v, date):
        return _correct_century_bug(v)
    if isinstance(v, (int, float)):
        return None  # not a real date
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
            try:
                return _correct_century_bug(datetime.strptime(v, fmt).date())
            except ValueError:
                continue
        return None
    return None


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"Missing source spreadsheet: {XLSX}")
    if not os.path.exists(SCHEMA_SQL):
        sys.exit(f"Missing schema file: {SCHEMA_SQL}")

    print(f"Source: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    print(f"Reading sheet '{ws.title}' ({ws.max_row} rows, {ws.max_column} cols)")

    # Connect, run schema, insert rows
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    print("Applying schema (drops and recreates cancelled_patent_research)...")
    with open(SCHEMA_SQL) as f:
        cur.execute(f.read())
    conn.commit()

    print("Reading and inserting rows...")
    rows_inserted = 0
    rows_skipped_blank = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            rows_skipped_blank += 1
            continue
        params = {
            "name":                  s(row[COL["name"]]),
            "allotment_number":      s(row[COL["allotment_number"]]),
            "tribe_reservation":     s(row[COL["tribe_reservation"]]),
            "state":                 s(row[COL["state"]]),
            "reason_for_cancellation": s(row[COL["reason_for_cancellation"]]),
            "cancellation_date":     d(row[COL["cancellation_date"]]),
            "fee_patent_date":       d(row[COL["fee_patent_date"]]),
            "patent_number":         s(row[COL["patent_number"]]),
            "ccf_number":            s(row[COL["ccf_number"]]),
            "gender":                s(row[COL["gender"]]),
            "carlisle_yn":           s(row[COL["carlisle_yn"]]),
            "comments":              s(row[COL["comments"]]),
            "in_dtpo":               s(row[COL["in_dtpo"]]),
            "ccf_alt":               s(row[COL["ccf_alt"]]),
            "source_row_index":      i,
        }
        cur.execute("""
            INSERT INTO cancelled_patent_research
              (name, allotment_number, tribe_reservation, state, reason_for_cancellation,
               cancellation_date, fee_patent_date, patent_number, ccf_number,
               gender, carlisle_yn, comments, in_dtpo, ccf_alt, source_row_index)
            VALUES
              (%(name)s, %(allotment_number)s, %(tribe_reservation)s, %(state)s,
               %(reason_for_cancellation)s, %(cancellation_date)s, %(fee_patent_date)s,
               %(patent_number)s, %(ccf_number)s, %(gender)s, %(carlisle_yn)s,
               %(comments)s, %(in_dtpo)s, %(ccf_alt)s, %(source_row_index)s)
        """, params)
        rows_inserted += 1

    conn.commit()
    print(f"  inserted: {rows_inserted}")
    print(f"  skipped (blank rows): {rows_skipped_blank}")

    # Verification: how many link cleanly to all_patents?
    cur.execute("""
        SELECT COUNT(*) AS n
        FROM cancelled_patent_research cpr
        JOIN all_patents ap ON ap.accession_number = cpr.patent_number
    """)
    linked = cur.fetchone()["n"]
    cur.execute("""
        SELECT COUNT(*) AS n
        FROM cancelled_patent_research
        WHERE patent_number IS NOT NULL
    """)
    with_pn = cur.fetchone()["n"]
    print()
    print(f"Records with patent_number populated: {with_pn}")
    print(f"  matching an all_patents row:        {linked}")
    print(f"  with patent_number but no DB match: {with_pn - linked}")

    # Of the linked ones, how many are flagged cancelled in all_patents?
    cur.execute("""
        SELECT
          SUM(CASE WHEN LOWER(ap.cancelled_doc::text) IN ('true','t') THEN 1 ELSE 0 END) AS flagged,
          COUNT(*) AS total
        FROM cancelled_patent_research cpr
        JOIN all_patents ap ON ap.accession_number = cpr.patent_number
    """)
    r = cur.fetchone()
    print(f"  of linked, BLM cancelled_doc=True:  {r['flagged']} / {r['total']}")

    # Reason distribution
    cur.execute("""
        SELECT reason_for_cancellation, COUNT(*) AS n
        FROM cancelled_patent_research
        WHERE reason_for_cancellation IS NOT NULL
        GROUP BY reason_for_cancellation
        ORDER BY n DESC
        LIMIT 20
    """)
    print()
    print("Reason distribution (top 20):")
    for row in cur.fetchall():
        print(f"  {row['n']:4d}  {row['reason_for_cancellation']}")

    # Sample of spreadsheet rows that have patent_number but DON'T match all_patents
    cur.execute("""
        SELECT cpr.patent_number, cpr.name, cpr.tribe_reservation, cpr.cancellation_date
        FROM cancelled_patent_research cpr
        LEFT JOIN all_patents ap ON ap.accession_number = cpr.patent_number
        WHERE cpr.patent_number IS NOT NULL AND ap.accession_number IS NULL
    """)
    unmatched = cur.fetchall()
    if unmatched:
        print()
        print(f"Spreadsheet entries with patent_number but no all_patents match ({len(unmatched)}):")
        for r in unmatched[:20]:
            print(f"  pn={r['patent_number']!r}  name={r['name']!r}  tribe={r['tribe_reservation']!r}")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
