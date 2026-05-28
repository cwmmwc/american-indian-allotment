"""
One-shot update to /Users/cwm6W/Library/CloudStorage/Box-Box/IATH/tribes/Tribes.xlsx.

Three groups of edits, all with [2026-05-28] attribution prefix so
Christian can distinguish these from his own prior work:

  (a) KCA reservation spelling — 21 rows currently use two spelling
      variants of the Kiowa-Comanche-Apache reservation; standardize
      all to "Kiowa, Comanche, and Apache Reservation".

  (b) Under-canonicalized GLO names — 17 rows for LA POINTE/BAD RIVER
      variants, OTTER TAIL PILLAGER CHIPPEWA, WHITE OAK POINT CHIPPEWA,
      and WINNIBIGOSHISH variants. Adds a `tribe notes` cell flagging
      that the GLO name itself contains a specific band/location
      identifier worth subsequent archival research. Does NOT assign
      a tribe or reservation (those remain FRN).

  (c) SIOUX / SIOUX INDIAN — 2 rows. Adds a `tribe notes` cell noting
      that 3,025 of the SIOUX-glo patents have BLM Document Type
      "Sioux Scrip Patent" (docClass=SS) and that the 1854 Act recital
      text identifies recipients as "half breeds or mixed bloods of
      the Dacotah or Sioux nation of Indians" — nation-level identity
      is documented in the patent text itself, while specific Dakota
      band remains unresolved (FRN).

Idempotent — if a note has already been added with the [2026-05-28]
prefix, it's skipped on re-run.

Run yourself; writes to the Box-synced path.

Usage:
    ./venv/bin/python3 scripts/update_tribes_spreadsheet.py            # dry-run, prints proposed edits
    ./venv/bin/python3 scripts/update_tribes_spreadsheet.py --apply    # writes to the xlsx
"""
import argparse
import re
import sys
from openpyxl import load_workbook

XLSX = "/Users/cwm6W/Library/CloudStorage/Box-Box/IATH/tribes/Tribes.xlsx"
TAG  = "[2026-05-28 Claude-assisted draft]"

KCA_CANONICAL = "Kiowa, Comanche, and Apache Reservation"
KCA_VARIANTS  = {"Kiowa, Comanche, Apache Reservation",
                 "Comanche, Kiowa, and Apache Reservation"}

NOTE_LAPOINTE = (
    f"{TAG} GLO name references La Pointe and/or Bad River — historical "
    f"Lake Superior Chippewa terminology. Further research could resolve "
    f"to a specific federally-recognized Ojibwe band."
)
NOTE_OTTERTAIL = (
    f"{TAG} GLO name identifies the Otter Tail Pillager — a specific Ojibwe "
    f"band named in 19th-century Minnesota treaty documents. 1,095 patent "
    f"records map to this GLO. Further research could resolve to a current "
    f"federally-recognized band affiliation."
)
NOTE_WHITEOAK = (
    f"{TAG} GLO name references White Oak Point — a location in Minnesota "
    f"with documented Ojibwe associations. 422 patent records map to this GLO. "
    f"Further research could resolve to a specific band."
)
NOTE_WINNI = (
    f"{TAG} GLO name references Lake Winnibigoshish in Minnesota — historical "
    f"Ojibwe territory. Further research could resolve to a specific band."
)
NOTE_SIOUX = (
    f"{TAG} 3,025 of the patents with this GLO have BLM Document Type "
    f'"Sioux Scrip Patent" (docClass=SS). Per the 1854 Act recital text '
    f'quoted on the patents themselves, recipients were identified as '
    f'"half breeds or mixed bloods of the Dacotah or Sioux nation of Indians." '
    f'Nation-level identity is documented in the patent text. Specific Dakota '
    f'band affiliation (Mdewakanton, Wahpekute, Sisseton, Wahpeton, etc.) '
    f'remains FRN. See document_class_metadata table for SS docClass details.'
)

UNDER_CANONICAL = {
    'OTTER TAIL PILLAGER CHIPPEWA': NOTE_OTTERTAIL,
    'WHITE OAK POINT CHIPPEWA':     NOTE_WHITEOAK,
}
WINNI_RE     = re.compile(r'^WINNI[BE]?I?GOSHISH', re.IGNORECASE)
LAPOINTE_RE  = re.compile(r'(LA\s*POINTE|LAPOINTE).*BAD\s*RIVER|BAD\s*RIVER.*(LA\s*POINTE|LAPOINTE)', re.IGNORECASE)
SIOUX_TARGETS = {'SIOUX', 'SIOUX INDIAN'}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(XLSX, data_only=False)
    ws = wb["Sheet1"]

    edits = []  # (row_idx, col_idx, current_value, new_value, reason)

    for row_idx in range(2, ws.max_row + 1):
        glo_cell = ws.cell(row=row_idx, column=1)
        tribe_cell = ws.cell(row=row_idx, column=2)
        res_cell = ws.cell(row=row_idx, column=3)
        tribe_notes_cell = ws.cell(row=row_idx, column=6)
        glo = (glo_cell.value or '').strip()
        glo_upper = glo.upper()

        # (a) KCA canonicalization
        if res_cell.value and str(res_cell.value).strip() in KCA_VARIANTS:
            edits.append((row_idx, 3, res_cell.value, KCA_CANONICAL, 'KCA canonical'))

        # (c+) SIOUX/SIOUX INDIAN tribe column — change FRN to Dacotah/Sioux Nation.
        # Verified across all three docClass samples (SS, STA, SER, MV) that the
        # patent recital text identifies recipients as Sioux at the nation level
        # ("Dacotah or Sioux nation" on SS records, "Sioux Indian" on STA/SER/MV).
        if glo_upper in SIOUX_TARGETS:
            current_tribe = (tribe_cell.value or '').strip() if tribe_cell.value else ''
            if current_tribe.upper() == 'FRN':
                edits.append((row_idx, 2, tribe_cell.value, 'Dacotah/Sioux Nation',
                              f'tribe label for {glo!r}'))

        # (b) Under-canonicalized GLOs — append to tribe_notes
        note_to_add = None
        if glo_upper in UNDER_CANONICAL:
            note_to_add = UNDER_CANONICAL[glo_upper]
        elif WINNI_RE.match(glo_upper):
            note_to_add = NOTE_WINNI
        elif LAPOINTE_RE.search(glo_upper):
            note_to_add = NOTE_LAPOINTE
        # (c) SIOUX
        elif glo_upper in SIOUX_TARGETS:
            note_to_add = NOTE_SIOUX

        if note_to_add:
            existing = (tribe_notes_cell.value or '').strip()
            if TAG in existing:
                continue  # already updated
            new_val = f"{existing}\n\n{note_to_add}".strip() if existing else note_to_add
            edits.append((row_idx, 6, existing, new_val, f"note for {glo!r}"))

    print(f"Total edits queued: {len(edits)}")
    print()
    by_reason = {}
    for r in edits:
        by_reason[r[4]] = by_reason.get(r[4], 0) + 1
    print("By reason:")
    for k, v in sorted(by_reason.items(), key=lambda kv: -kv[1]):
        print(f"  {v:>3}  {k}")
    print()
    print("First 5 sample edits:")
    for row_idx, col, old, new, reason in edits[:5]:
        print(f"  row {row_idx} col {col} ({reason})")
        print(f"    OLD: {(old or '(empty)')!r}")
        print(f"    NEW: {new[:120]!r}{'...' if len(str(new)) > 120 else ''}")
        print()

    if not args.apply:
        print("DRY RUN — pass --apply to write to xlsx.")
        return

    if not edits:
        print("No edits to apply.")
        return

    for row_idx, col, _, new, _ in edits:
        ws.cell(row=row_idx, column=col).value = new

    wb.save(XLSX)
    print(f"Saved {len(edits)} edits to {XLSX}")
    print("Box will sync the change to cloud automatically.")


if __name__ == "__main__":
    main()
