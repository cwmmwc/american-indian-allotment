"""
One-shot: normalize 5 casing/whitespace artifacts in the IATH Tribes
spreadsheet, column B ("Authoritative Tribe Name"). Per Christian's
2026-06-03 direction:

  ABSENTEE SHAWNEE     -> Absentee Shawnee
  ABSENTEE WYANDOTTE   -> Absentee Wyandotte
  CITIZEN POTAWATOMI   -> Citizen Potawatomi
  OTOE AND MISSOURIA   -> Otoe and Missouria
  "FRN " (trailing sp) -> FRN

Default is dry-run: prints every cell that would change and exits
without touching the file. Run with --write to apply, which:
  1. Makes a timestamped .bak copy alongside the original.
  2. Writes the normalized workbook in place.

The script only touches cells in column B whose value is an EXACT match
for one of the five source strings above. Other columns and other
values are not modified. FRN (no trailing space) is left alone.

Refuses to run if column B's header isn't "Authoritative Tribe Name"
(safety against the spreadsheet schema having changed).
"""
import argparse
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

SPREADSHEET = Path(
    "/Users/cwm6W/Library/CloudStorage/Box-Box/IATH/tribes/Tribes.xlsx"
)
EXPECTED_COL_B_HEADER = "Authoritative Tribe Name"

# Exact cell value -> canonical replacement.
TRANSFORMS = {
    "ABSENTEE SHAWNEE":   "Absentee Shawnee",
    "ABSENTEE WYANDOTTE": "Absentee Wyandotte",
    "CITIZEN POTAWATOMI": "Citizen Potawatomi",
    "OTOE AND MISSOURIA": "Otoe and Missouria",
    "FRN ":               "FRN",
}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--write", action="store_true",
        help="Apply changes in place (default: dry-run only).",
    )
    args = ap.parse_args()

    if not SPREADSHEET.exists():
        raise SystemExit(f"not found: {SPREADSHEET}")

    wb = openpyxl.load_workbook(SPREADSHEET)
    ws = wb.active

    header = ws.cell(row=1, column=2).value
    if header != EXPECTED_COL_B_HEADER:
        raise SystemExit(
            f"column B header is {header!r}, expected "
            f"{EXPECTED_COL_B_HEADER!r} — refusing to edit."
        )

    changes = []
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        v = cell.value
        if v in TRANSFORMS:
            glo = ws.cell(row=row, column=1).value or ""
            changes.append((row, glo, v, TRANSFORMS[v]))

    print(f"Spreadsheet:  {SPREADSHEET}")
    print(f"Active sheet: {ws.title!r}, header row matches expected.")
    print(f"Cells matching a transform: {len(changes)}")
    print()
    print(f"{'row':>5}  {'Name from GLO':<36}  {'before':<22}  ->  after")
    print(f"{'-' * 5}  {'-' * 36}  {'-' * 22}      {'-' * 22}")
    for row_idx, glo, before, after in changes:
        glo_str = (str(glo) or "")[:36]
        print(f"{row_idx:>5}  {glo_str:<36}  {before!r:<22}  ->  {after!r}")

    if not args.write:
        print()
        print("DRY RUN — nothing written. Re-run with --write to apply.")
        return

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = SPREADSHEET.with_suffix(f".bak_{ts}.xlsx")
    shutil.copy2(SPREADSHEET, bak)
    print(f"\nBackup written: {bak}")

    for row_idx, _, _, after in changes:
        ws.cell(row=row_idx, column=2).value = after
    wb.save(SPREADSHEET)
    print(f"Applied {len(changes)} normalizations to {SPREADSHEET}")


if __name__ == "__main__":
    main()
